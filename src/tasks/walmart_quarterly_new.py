#!/usr/bin/env python3

import io
from datetime import datetime as dt
from decimal import Decimal
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from hippo import logger
from hippo.sources.s3 import S3Downloader
from sources import claims as sources_claims
from sources import pharmacies
from tasks import download_other_sources
from transform import claims as transform_claims
from transform import utils
from transform.walmart import walmart_claims
from . import invoice_utils as iu

log = logger.getLogger('Processing Walmart quarterly') 

chain_name = 'walmart'
chain_class = pharmacies.Walmart

def run(invoice_bucket, slack_channel, slack_bot_token, hsdk_env, period, manual_date_string, **kwargs):

    period_start, period_end = utils.process_invoice_dates(period_flag='quarter')
    # Walmart wants to receive the data starting the Contract Effective Date (2024-09-03) if reporting quarter is in 2024 or January 1st for the next years
    # we already exclude all claims with claim_date_of_service >= '2024-09-03' and out of month reversals with valid_from < '2024-09-03' which means we only need to implement solution starting q1 2025
    current_year = period_start.year
    if current_year != 2024:
        period_start = dt.strptime(f'{current_year}-01-01', "%Y-%m-%d")
    if current_year == 2024:
        period_start = dt.strptime(f'2024-09-03', "%Y-%m-%d")

    log.info(f"Running {period}ly report.")
    log.info(f"Running date is {manual_date_string}.")
    temp_main_prefix = f"temp/sources/walmart/{period}/"
    temp_chain_specific_prefix = f"temp/sources/specific/walmart/{period}/"
    temp_claims_prefix = f"temp/claims/walmart/{period}"
    
                                        ### DOWNLOAD AND PROCESS CLAIMS AND SOURCE FILES ###
    log.info(f"Downloading the data for the period: {period_start} - {period_end}.")
    raw_claims_df = S3Downloader(invoice_bucket, temp_claims_prefix, 'claims.parquet').pull_parquet()[0]
    claims_df = _preprocess_claims_df(raw_claims_df, period_start, period_end) # apply all necessary filters and conditions for specific chain
    log.info(f"Processing default source dictionary.")
    source_dict = download_other_sources.load_parquet_files_to_dict(invoice_bucket, temp_main_prefix)
    log.info(f"Processing CVS specific source dictionary.")
    walmart_source_dict = download_other_sources.load_parquet_files_to_dict(invoice_bucket, temp_chain_specific_prefix)


    claims_df = walmart_claims.join_ndcs_sources_for_specific_date(claims_df, walmart_source_dict)
    claims_df = walmart_claims.join_excluded_ndcs(claims_df, walmart_source_dict['walmart_excluded'])
    claims_df = walmart_claims.join_four_dollar_ndcs(claims_df, walmart_source_dict['walmart_$4_ndcs'])

                                        ### PROCESS TRANSACTIONS ###
    transaction_raw_df = transform_claims.build_default_transactions_dataset(claims_df, source_dict, chain_name, chain_class, period_start, period_end)
    transaction_walmart_df = walmart_claims.add_specific_walmart_columns(transaction_raw_df)
    transaction_final_df = _select_fields_for_exporting(transaction_walmart_df)

                                            ### VALIDATE PRICE BASIS ###
    # need this validation in order to ensure that we don't blindly reconcile agains reconciliation_price_basis
    # if any descripancies found - we will receive notification in #temp_invoices slack channel and decide is it expected or not
    validation_dict = utils.validate_price_basis(transaction_raw_df, period_start, period_end, invoice_bucket, chain_name)
    if validation_dict['dataframe'].empty == False:
        utils.send_text_message_to_slack(kwargs.get('temp_slack_channel'), slack_bot_token, validation_dict['message'])

                                        ### PROCESS SUMARY REPORT ###
    ger_results = walmart_claims.build_walmart_quarterly_summary(transaction_raw_df)

                                        ### PROCESS EXCLUDED ITEMS REPORT ###
    excluded_items_report_df = walmart_claims.build_excluded_items_dataframe(transaction_raw_df)

                                        ### PROCESS ADMIN FEE REPORT ###
    admin_fee_report_df = walmart_claims.admin_fee_report_dataframe(transaction_raw_df)

                                        ### VALIDATE RESULTS: SUMMARY VS TRANSACTION FILE ###
    _validate_walmart_results(excluded_items_report_df, admin_fee_report_df, ger_results, transaction_final_df)

    result_dict = {'Excluded Items Reconciled at U&C': excluded_items_report_df, 'Discount Card Admin Fee Reconciliation': admin_fee_report_df, 'Discount Card': ger_results}

    excel_file = export_dict_to_excel(result_dict)

    transaction_file_name = utils.transactions_file_name(chain_name=chain_name.upper(), report_type='quarterly', period_start=period_start.strftime('%Y-%m-%d')+'_')
    summary_file_name = utils.transactions_file_name(chain_name=chain_name.upper(), report_type='summary', period_start=period_start.strftime('%Y-%m-%d')+'_')

    iu.send_file([ger_results, excluded_items_report_df, admin_fee_report_df], chain_name.upper(), summary_file_name, excel_file, slack_channel, slack_bot_token, invoice_bucket, hsdk_env=hsdk_env, **kwargs)
    iu.create_and_send_file("Utilization", transaction_final_df, transaction_file_name, slack_channel, slack_bot_token, invoice_bucket, chain=chain_name.upper(), hsdk_env=hsdk_env, **kwargs)


def _preprocess_claims_df(raw_claims_df, period_start, period_end):
    claims_df = raw_claims_df.drop_duplicates(subset=['authorization_number'], keep='first')
    claims_df = sources_claims.preprocess_claims(claims_df, partners=True)
    claims_df = claims_df.drop_duplicates(subset=['authorization_number'], keep='first')                                      # in rare cases we have duplicated claims because of duplcates in reporting.cardholder cases; in this particular case it's not so important what partner will be joined so we can keep first
    claims_df = transform_claims.filter_by_binnumber_authnumber(claims_df, chain_class.BIN_NUMBERS)                           # filter dataframe by bin_number and authorization_number
    claims_df = claims_df[claims_df['claim_date_of_service'].dt.floor('D') >= chain_class.CONTRACT_START_DATE]                # go live for Walmart new contract 2024-09-03
    claims_df = sources_claims.filter_reversals(claims_df, period_start, period_end)
    return claims_df

def _select_fields_for_exporting(transactions_df):
    transactions_df['NDC'] = transactions_df['product_id'].astype('string').str.zfill(11)
    transactions_df['claim_date_of_service'] = transactions_df['claim_date_of_service'].apply(lambda x: x.strftime('%Y-%m-%d')).fillna('')
    transactions_df = transactions_df.rename(columns=chain_class.COLUMN_RENAME_DICT) # rename columns
    requested_columns = set(chain_class.REQUESTED_COLUMNS) - set(['Reversal date', 'Fill/Reversal indicator'])
    requested_columns = list(requested_columns)
    transactions_df = transactions_df[requested_columns]
    return transactions_df

def export_dict_to_excel(result_dict):
    ACCOUNTING_FORMAT = '"$"#,##0.00_);("$"#,##0.00)'

    wb = Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    header_font = Font(bold=True, color='000000', size=14)  # Black font, bold, larger size
    header_fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')  # Light blue background
    column_header_font = Font(bold=True, color='000000')  # Black bold font for column headers
    column_fill = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')  # Light grey background for column headers and last row
    last_row_fill = PatternFill(start_color='00C0C0C0', end_color='00C0C0C0', fill_type='solid')  # Light grey background for the last row

    # List of columns that require specific formatting
    accounting_columns = [
        'Amt Under U&C', 'Full Cost', 'Ingredient Cost Paid', 'Contracted Dispensig Fee',
        'Dispensing Fee Paid', 'Rate Variance', 'DF Variance', 'Total Variance', 
        'Admin Fee Paid', 'Variance'
    ]
    float_columns = ['Contract Rate', 'Actual Rate', 'Contracted Admin', 'Admin Fee /Rx', 'Delta']
    integer_columns = ['Claim Count']

    # Function to format cells
    def format_cells(ws, range_ref, alignment='center', font=None, border=True, fill=None):
        for row in ws[range_ref]:
            for cell in row:
                cell.alignment = Alignment(horizontal=alignment)
                if font:
                    cell.font = font
                if border:
                    cell.border = thin_border
                if fill:
                    cell.fill = fill

    row_offset = 1
    for key, df in result_dict.items():
        # Merged cell for the header with light blue background, left aligned
        col_count = len(df.columns)
        ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=col_count)
        header_cell = ws.cell(row=row_offset, column=1, value=key)
        header_cell.font = header_font
        header_cell.fill = header_fill
        format_cells(ws, f'A{row_offset}:{chr(64 + col_count)}{row_offset}', alignment='left')

        row_offset += 1

        # Add column headers with the same formatting as the last row
        for col_num, col_name in enumerate(df.columns, 1):
            ws.cell(row=row_offset, column=col_num, value=col_name).font = column_header_font
        format_cells(ws, f'A{row_offset}:{chr(64 + col_count)}{row_offset}', fill=column_fill)

        row_offset += 1

        # Add DataFrame rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), row_offset):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if df.columns[c_idx - 1] in accounting_columns:
                    cell.number_format = ACCOUNTING_FORMAT  # Use custom accounting format
                elif df.columns[c_idx - 1] in float_columns:
                    cell.number_format = '0.000'
                elif df.columns[c_idx - 1] in integer_columns:
                    cell.number_format = '0'
        format_cells(ws, f'A{row_offset}:{chr(64 + col_count)}{row_offset + len(df) - 1}')

        # Last row formatting (Total row) with light grey background
        last_row_idx = r_idx
        format_cells(ws, f'A{last_row_idx}:{chr(64 + col_count)}{last_row_idx}', fill=last_row_fill, font=column_header_font)

        # If it's the 'Discount Card' DataFrame, adjust the column widths
        if key == 'Discount Card':
            for col_num, col_name in enumerate(df.columns, 1):
                if col_num == 1:
                    # For the first column, set width to 3 times the length of the column name
                    ws.column_dimensions[chr(64 + col_num)].width = len(col_name) * 3
                else:
                    # For other columns, set width to the length of the column name
                    ws.column_dimensions[chr(64 + col_num)].width = len(col_name) + 2  # Add padding to prevent truncation

        # Insert a blank row before the next DataFrame
        row_offset = r_idx + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def _validate_walmart_results(excluded_items_report_df, admin_fee_report_df, ger_results, transaction_final_df):
    transactions_df = transaction_final_df[
        ['Total Quantity Dispensed', '30/90 Day Claim Indicator', 'Brand/Generic Designation', 'Contracted Rate',
         'Contracted Dispense Fee', 'Reimbursement Type', 'Actual Rate', 'AWP', 'NADAC', 'GPI_NADAC',
         'Ingredient Cost Paid', 'Dispensing Fee Paid', 'Administration Fee', 'Variant Amount',
         'Excluded Claim Indicator', 'U&C Amount', 'Patient Pay Response', 'Taxes Paid']
    ]

    excluded_items_report_df = excluded_items_report_df[excluded_items_report_df['Exclusion Category'] != 'Total']
    admin_fee_report_df = admin_fee_report_df[admin_fee_report_df['Network'] != 'Total']
    ger_results = ger_results[ger_results['Network'] != 'Total']

    ### Validate Excluded Items Block
    excluded_df = transactions_df[transactions_df['Excluded Claim Indicator'] == 'Y']
    total_excluded_count = excluded_df.shape[0]
    total_excluded_amt = (excluded_df['U&C Amount'] - (excluded_df['Patient Pay Response'] - excluded_df['Taxes Paid'])).sum()

    assert excluded_items_report_df['Claim Count'].sum() == total_excluded_count, \
        f"Excluded claims validation failed: Claim Count total from Summary {excluded_items_report_df['Claim Count'].sum()} != Utilization {total_excluded_count}"

    assert excluded_items_report_df['Amt Under U&C'].sum() == total_excluded_amt, \
        f"Excluded claims validation failed: Amt Under U&C from Summary {excluded_items_report_df['Amt Under U&C'].sum()} != Utilization {total_excluded_amt}"

    ### Validate Admin Fee Block
    admin_fee_df = transactions_df[transactions_df['Excluded Claim Indicator'] == 'N']
    total_admin_fee_count = admin_fee_df.shape[0]
    total_admin_fee_paid = admin_fee_df['Administration Fee'].sum()
    total_admin_fee_variance = admin_fee_df['Administration Fee'].sum() - Decimal(chain_class.ADMIN_FEE).quantize(Decimal('0.01')) * total_admin_fee_count

    assert admin_fee_report_df['Rx Count'].sum() == total_admin_fee_count, \
        f"Admin Fee validation failed: Summary Claim Count {admin_fee_report_df['Rx Count'].sum()} != Utilization {total_admin_fee_count}"

    assert admin_fee_report_df['Admin Fee Paid'].sum() == total_admin_fee_paid, \
        f"Admin Fee validation failed: Administration Fee Paid in Summary {admin_fee_report_df['Admin Fee Paid'].sum()} != Utilization {total_admin_fee_paid}"

    assert abs(admin_fee_report_df['Variance'].sum() - total_admin_fee_variance) <= Decimal('1'), \
        f"Admin Fee validation failed: Admin Fee Variance {admin_fee_report_df['Variance'].sum()} != Utilization {total_admin_fee_variance}"

    ### Validate Discount Card Block
    discount_card_df = transactions_df[transactions_df['Excluded Claim Indicator'] == 'N']
    discount_card_count = discount_card_df.shape[0]
    discount_card_ic_sum = discount_card_df['Ingredient Cost Paid'].sum()
    discount_card_df_sum = discount_card_df['Dispensing Fee Paid'].sum()
    discount_card_variance_sum = discount_card_df['Variant Amount'].sum()

    assert ger_results['Claim Count'].sum() == discount_card_count, \
        f"Discount Card summary validation failed: Summary Claim Count {ger_results['Claim Count'].sum()} != Utilization {discount_card_count}"

    assert ger_results['Ingredient Cost Paid'].sum() == discount_card_ic_sum, \
        f"Discount Card summary validation failed: Ingredient Cost in Summary {ger_results['Ingredient Cost Paid'].sum()} != Utilization {discount_card_ic_sum}"

    assert ger_results['Dispensing Fee Paid'].sum() == discount_card_df_sum, \
        f"Discount Card summary validation failed: Dispensing Fee in Summary {ger_results['Dispensing Fee Paid'].sum()} != Utilization {discount_card_df_sum}"

    assert abs(ger_results['Total Variance'].sum() - discount_card_variance_sum) <= Decimal('1'), \
        f"Discount Card summary validation failed: Total Variance in Summary {ger_results['Total Variance'].sum()} != Utilization {discount_card_variance_sum}"
