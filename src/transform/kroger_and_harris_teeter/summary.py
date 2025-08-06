import pandas as pd
from decimal import Decimal
from sources import pharmacies
from sources import claims as sources_claims
from transform import claims, utils

from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

chain = pharmacies.KrogerAndHarrisTeeter
DISP_FEE_DICT = pharmacies.KrogerAndHarrisTeeter.DISP_FEE_DICT

def transform_data(raw_claims_df, source_dict, period_start, period_end, summary_period):
    summary_df = sources_claims.join_default_source_dataframes(raw_claims_df, source_dict)

    summary_df['claim_type'] = chain.determine_dea_claim_type(summary_df['dea_class_code'])
    summary_df['total_paid_response'] = -summary_df['total_paid_response']
    summary_df['fill_type'] = claims.determine_basis_of_rembursment_abbreviation(summary_df['basis_of_reimbursement_determination_resp'])
    summary_df['days_supply'] = _process_days_supply(summary_df['days_supply'])
    summary_df['Unique Identifier'] = summary_df['bin_number'].astype(str) + summary_df['process_control_number'].astype(str) + summary_df['network_reimbursement_id'].astype(str)
    summary_df['drug_type'] = chain.add_brand_generic_indicator(summary_df['multi_source_code'])

    result_dict = {}
    # process current period SUMMARY data
    current_period_summary_df = summary_df[(summary_df.valid_from.dt.floor('D') >= period_start) & (summary_df.valid_from.dt.floor('D') <= period_end)] # filter out the claim outside of the current month

    for unique_identifier in current_period_summary_df['Unique Identifier'].unique():
        grouped_df_dict = _group_up_the_data(current_period_summary_df[current_period_summary_df['Unique Identifier'] == unique_identifier])
        
        for drug_type in ['brand', 'generic']:
            if drug_type in grouped_df_dict:  # Process only if the key exists
                grouped_df_dict[drug_type] = utils.cast_columns_to_decimal(
                    grouped_df_dict[drug_type],
                    column_names=[
                        'Actual NADAC ingredient cost paid ($)',
                        'Contracted NADAC ingredient cost paid ($)',
                        'Actual MAC ingredient cost paid',
                        'Contracted MAC ingredient cost paid ($)',
                        'Actual AWP ingredient cost discount',
                        'Total Administration Fees ($)',
                        'Dispensing Fees paid ($)',
                        'Total AWP ($)',
                        'Actual AFER ($)'
                    ],
                    fillna_flag=True
                )
                grouped_df_dict[drug_type] = utils.cast_cents_to_dollars(
                    grouped_df_dict[drug_type],
                    column_names=[
                        'Actual NADAC ingredient cost paid ($)',
                        'Contracted NADAC ingredient cost paid ($)',
                        'Actual MAC ingredient cost paid',
                        'Contracted MAC ingredient cost paid ($)',
                        'Total Administration Fees ($)',
                        'Dispensing Fees paid ($)'
                    ],
                    add_fill_reversal_sign=False
                )
        
        # Add the processed dictionary to the result
        result_dict[unique_identifier] = grouped_df_dict
    
    # add current month total data
    result_dict = _group_up_total_summary(result_dict, 'current-period-total')

    # process year-to-date SUMMARY data
    result_dict = _group_up_total_year_summary(summary_df, result_dict, 'year-to-date-total')

    # create excel
    result_bytes = create_excel_from_dict(result_dict, period_start, period_end, summary_period)
    result_dataframe = pd.read_excel(result_bytes, sheet_name='Sheet')
    return result_bytes, result_dataframe


def _group_up_the_data(df):
    grouped_df_dict = {}
    for drug_type in ['brand','generic']:

        filtered_df = df[(df.fill_type != 'UNC') & (df['drug_type'] == drug_type)]

        if not filtered_df.empty:

            try:
                result_df = filtered_df.groupby(
                    ['Unique Identifier', 'claim_type', 'days_supply', 'bin_number', 'process_control_number',
                     'drug_type', 'multi_source_code', 'network_reimbursement_id']).agg(  # UNC excluded
                    Total_number_of_paid_claims=pd.NamedAgg(column='fill_type', aggfunc='size'),
                    Total_number_of_NADAC_paid_claims=pd.NamedAgg(column='fill_type',
                                                                  aggfunc=lambda x: (x == 'NADAC').sum()),
                    Actual_NADAC_ingredient_cost_paid=pd.NamedAgg(column='ingredient_cost_paid_resp',
                                                                  aggfunc=lambda x: x[
                                                                      filtered_df['fill_type'] == 'NADAC'].astype(float).sum()),

                    Contracted_NADAC_ingredient_cost_paid=pd.NamedAgg(column='fill_type', aggfunc=lambda x: (((filtered_df.loc[
                                                                                                                   x.index, 'nadac'].astype(
                        float) * filtered_df.loc[x.index, 'quantity_dispensed'].astype(float)).round(0))[
                        filtered_df.loc[x.index, 'fill_type'] == 'NADAC']).sum()),
                    Total_number_of_AWP_paid_claims=pd.NamedAgg(column='fill_type',
                                                                aggfunc=lambda x: (x == 'AWP').sum()),
                    Total_AWP=pd.NamedAgg(column='fill_type', aggfunc=lambda x: (
                    ((filtered_df.loc[x.index, 'awp'].astype(float) * filtered_df.loc[x.index, 'quantity_dispensed'].astype(float)))[
                        filtered_df.loc[x.index, 'fill_type'] == 'AWP']).sum()),
                    Contracted_AWP_ingredient_cost_paid=pd.NamedAgg(column='ingredient_cost_paid_resp',
                                                                    aggfunc=lambda x: x[
                                                                        filtered_df['fill_type'] == 'AWP'].astype(float).sum()),
                    Total_number_of_MAC_paid_claims=pd.NamedAgg(column='fill_type',
                                                                aggfunc=lambda x: (x == 'MAC').sum()),
                    Total_MAC=pd.NamedAgg(column='fill_type', aggfunc=lambda x: (
                    ((filtered_df.loc[x.index, 'mac'].astype(float) * filtered_df.loc[x.index, 'quantity_dispensed'].astype(float)))[
                        filtered_df.loc[x.index, 'fill_type'] == 'MAC']).sum()),
                    Actual_MAC_ingredient_cost_paid=pd.NamedAgg(column='ingredient_cost_paid_resp',
                                                                    aggfunc=lambda x: x[
                                                                        filtered_df['fill_type'] == 'MAC'].astype(float).sum()),
                    Contracted_MAC_ingredient_cost_paid=pd.NamedAgg(column='fill_type', aggfunc=lambda x: (((filtered_df.loc[
                                                                                                                   x.index, 'mac'].astype(
                        float) * filtered_df.loc[x.index, 'quantity_dispensed'].astype(float)).round(0))[
                        filtered_df.loc[x.index, 'fill_type'] == 'MAC']).sum()),
                    Dispensing_Fees_paid=pd.NamedAgg(column='dispensing_fee_paid_resp', aggfunc='sum'),
                    Total_Administration_Fees=pd.NamedAgg(column='total_paid_response', aggfunc='sum'),
                    Actual_AFER=pd.NamedAgg(column='total_paid_response', aggfunc='mean'),
                ).reset_index()

            except Exception as exc:
                raise exc

            result_df['Actual_AWP_ingredient_cost_discount'] = (
                    (result_df['Total_AWP'] - result_df['Contracted_AWP_ingredient_cost_paid']) / result_df['Total_AWP']
                ).round(3)

            result_df['Total_AWP'] = result_df['Total_AWP'].divide(100).round(2)
            result_df['Total_MAC'] = result_df['Total_MAC'].divide(100).round(2)
            result_df['Actual_AFER'] = result_df['Actual_AFER'].divide(100).round(4)
            result_df['Targeted_AFER'] = float(chain.TARGET_AFER)
            result_df['Contracted AWP Ingredient Cost AWP discount'] = float(chain.TARGET_BER) if drug_type == 'brand' else float(chain.TARGET_GER)
            result_df['AFER Diff ($)'] = -((result_df['Actual_AFER'] - result_df['Targeted_AFER']) * result_df['Total_number_of_paid_claims']).round(2)
            result_df['NADAC Diff ($)'] = -(result_df['Actual_NADAC_ingredient_cost_paid'].divide(100) - result_df['Contracted_NADAC_ingredient_cost_paid'].divide(100)).round(2)
            result_df['AWP Diff ($)'] = ((result_df['Contracted AWP Ingredient Cost AWP discount'] - result_df['Actual_AWP_ingredient_cost_discount']) * result_df['Total_AWP']).round(2).fillna(0)

            result_df['MAC Diff ($)'] = -(result_df['Actual_MAC_ingredient_cost_paid'].divide(100) - result_df['Contracted_MAC_ingredient_cost_paid'].divide(100)).round(2)

            result_df['Ingredient Cost Diff ($)'] = ((result_df['NADAC Diff ($)'] + result_df['AWP Diff ($)'] + result_df['MAC Diff ($)'])).round(2).fillna(0)


            result_df = calculate_dispensing_fee_total(result_df, drug_type)
            result_df['Dispensing Fee Diff ($)'] = -(result_df['Dispensing_Fees_paid'].divide(100) - result_df['dispensing_fee_total'])
            result_df = result_df.rename(columns={
                'network_reimbursement_id': 'Network ID',
                'process_control_number': 'PCN',
                'bin_number': 'BIN',
                'multi_source_code':'MONY code',
                'claim_type': 'Claim type',
                'days_supply': 'Days supply',
                'Total_number_of_paid_claims': 'Total number of paid claims',
                'Total_number_of_NADAC_paid_claims': 'Total number of NADAC paid claims',
                'Actual_NADAC_ingredient_cost_paid': 'Actual NADAC ingredient cost paid ($)',
                'Contracted_NADAC_ingredient_cost_paid': 'Contracted NADAC ingredient cost paid ($)',
                'Total_number_of_AWP_paid_claims': 'Total number of AWP paid claims',
                'Total_AWP': 'Total AWP ($)',
                'Actual_AWP_ingredient_cost_discount': 'Actual AWP ingredient cost discount',
                'Total_number_of_MAC_paid_claims': 'Total number of MAC paid claims',
                'Total_MAC': 'Total MAC ($)',
                'Contracted_MAC_ingredient_cost_paid': 'Contracted MAC ingredient cost paid ($)',
                'Actual_MAC_ingredient_cost_paid': 'Actual MAC ingredient cost paid',
                'Dispensing_Fees_paid': 'Dispensing Fees paid ($)',
                'Total_Administration_Fees': 'Total Administration Fees ($)',
                'Actual_AFER': 'Actual AFER ($)',
                'Targeted_AFER': 'Targeted AFER ($)',
                'drug_type': 'Drug type'
                })

            result_df = result_df[['Unique Identifier', 'Claim type', 'Days supply', 'BIN', 'PCN', 'Network ID', 'Drug type', 'MONY code', 'Total number of paid claims',
                'Total number of NADAC paid claims', 'Actual NADAC ingredient cost paid ($)',
                'Contracted NADAC ingredient cost paid ($)', 'NADAC Diff ($)', 'Total number of AWP paid claims', 'Total AWP ($)',
                'Contracted AWP Ingredient Cost AWP discount', 'Actual AWP ingredient cost discount', 'AWP Diff ($)',
                'Total number of MAC paid claims', 'Total MAC ($)', 'Contracted MAC ingredient cost paid ($)', 'Actual MAC ingredient cost paid', 'MAC Diff ($)',
                'Dispensing Fees paid ($)', 'Total Administration Fees ($)', 'Actual AFER ($)', 'Targeted AFER ($)', 'AFER Diff ($)', 'Ingredient Cost Diff ($)', 'Dispensing Fee Diff ($)']]
            grouped_df_dict[drug_type] = result_df[result_df['Total number of paid claims'] != 0]
    return grouped_df_dict


def calculate_dispensing_fee_total(df, drug_type):
    df['dispensing_fee_total'] = Decimal(0)
    
    for idx, row in df.iterrows():
        claim_type = row['claim_type']
        days_supply = row['days_supply']
        total_nadac_claims = row['Total_number_of_NADAC_paid_claims']
        total_awp_claims = row['Total_number_of_AWP_paid_claims']
        total_mac_claims = row['Total_number_of_MAC_paid_claims']

        if claim_type in DISP_FEE_DICT:
            # Calculate NADAC / MAC dispensing fee
            nadac_fee = Decimal(0)
            mac_fee = Decimal(0)
            awp_fee = Decimal(0)

            disp_fee = Decimal(DISP_FEE_DICT[claim_type][drug_type][days_supply])
            if days_supply in DISP_FEE_DICT[claim_type][drug_type]:
                nadac_fee = disp_fee * Decimal(total_nadac_claims)
                mac_fee = disp_fee * Decimal(total_mac_claims)

            # Calculate AWP dispensing fee
            if days_supply in DISP_FEE_DICT['AWP'][drug_type]:
                awp_fee = Decimal(DISP_FEE_DICT['AWP'][drug_type][days_supply]) * Decimal(total_awp_claims)
            
            # Sum up the fees
            df.at[idx, 'dispensing_fee_total'] += nadac_fee + awp_fee + mac_fee

    return df

def _process_days_supply(days_supply_series: pd.Series) -> pd.Series:
    return days_supply_series.apply(lambda x: '1-83' if x <= 83 else '84+')

def _group_up_total_year_summary(df, grouped_df_dict, period):
    filtered_df = df[df.fill_type != 'UNC']
    
    grouped_df = filtered_df.groupby(lambda _: True).agg({
        'fill_type': ['size', lambda x: (x == 'NADAC').sum(), lambda x: (x == 'AWP').sum(), lambda x: (x == 'MAC').sum()],
        'ingredient_cost_paid_resp': [lambda x: x[filtered_df['fill_type'] == 'NADAC'].astype(float).sum(),
                                      lambda x: x[filtered_df['fill_type'] == 'AWP'].astype(float).sum(),
                                      lambda x: x[filtered_df['fill_type'] == 'MAC'].astype(float).sum()
                                      ],
        'nadac': [lambda x: (((x.astype(float) * filtered_df['quantity_dispensed'].astype(float)).round(0))[filtered_df['fill_type'] == 'NADAC']).sum()],
        'awp': [lambda x: (((x.astype(float) * filtered_df['quantity_dispensed'].astype(float)))[filtered_df['fill_type'] == 'AWP']).sum()],
        'mac': [lambda x: (((x.astype(float) * filtered_df['quantity_dispensed'].astype(float)))[filtered_df['fill_type'] == 'MAC']).sum()],
        'dispensing_fee_paid_resp': 'sum',
        'total_paid_response': ['sum', 'mean']
    })

    grouped_df.columns = [
        'Total_number_of_paid_claims', 
        'Total_number_of_NADAC_paid_claims', 
        'Total_number_of_AWP_paid_claims',
        'Total_number_of_MAC_paid_claims',
        'Actual_NADAC_ingredient_cost_paid',
        'Contracted_AWP_ingredient_cost_paid',
        'Actual_MAC_ingredient_cost_paid',
        'Contracted_MAC_ingredient_cost_paid',
        'Total_AWP',
        'Total_MAC',
        'Dispensing_Fees_paid',
        'Total_Administration_Fees',
        'Actual_AFER'
    ]

    # Adding calculated fields
    grouped_df['Actual_AWP_ingredient_cost_discount'] = (
        (grouped_df['Total_AWP'] - grouped_df['Contracted_AWP_ingredient_cost_paid']) / grouped_df['Total_AWP']
    ).divide(100).round(4)

    grouped_df['Total_AWP'] = grouped_df['Total_AWP'].divide(100).round(2)
    grouped_df['Actual_AFER'] = grouped_df['Actual_AFER'].divide(100).round(3)
    grouped_df = utils.cast_columns_to_decimal(grouped_df, column_names=['Actual_NADAC_ingredient_cost_paid',
    'Contracted_NADAC_ingredient_cost_paid', 'Actual_AWP_ingredient_cost_discount', 'Total_Administration_Fees',
    'Dispensing_Fees_paid', 'Contracted_AWP_ingredient_cost_paid','Total_AWP', 'Contracted_MAC_ingredient_cost_paid',
    'Actual_MAC_ingredient_cost_paid', 'Dispensing_Fees_paid', 'Actual_AFER'], fillna_flag = True)
    grouped_df = utils.cast_cents_to_dollars(grouped_df, column_names=['Actual_NADAC_ingredient_cost_paid', 'Contracted_NADAC_ingredient_cost_paid', 'Actual_MAC_ingredient_cost_paid', 'Contracted_MAC_ingredient_cost_paid','Total_Administration_Fees', 'Dispensing_Fees_paid', 'Contracted_AWP_ingredient_cost_paid'], add_fill_reversal_sign=False)
    grouped_df = grouped_df[['Total_number_of_paid_claims', 'Total_number_of_NADAC_paid_claims', 'Total_number_of_MAC_paid_claims',
                             'Total_number_of_AWP_paid_claims', 'Actual_AFER']]
    grouped_df_dict[period] = grouped_df
    
    return grouped_df_dict

def _group_up_total_summary(result_dict, period):

    grand_totals = {
        "AFER Diff ($)": 0.0,
        "Ingredient Cost Diff ($)": 0.0,
        "Dispensing Fee Diff ($)": Decimal(0)
    }
    
    # Loop through each unique combination in the dictionary and sum up Diff. The combinations are: 1st dimention key = Unique_identificatr --> 2nd dimention keys are 'brand' or 'generic'
    for unique_combination, drug_types in result_dict.items():
        # Loop through 'brand' and 'generic'
        for drug_type, df in drug_types.items():
            # Check if the dataframe is not empty
            if not df.empty:
                # Sum the columns of interest
                grand_totals["AFER Diff ($)"] += df["AFER Diff ($)"].sum()
                grand_totals["Ingredient Cost Diff ($)"] += df["Ingredient Cost Diff ($)"].sum()
                grand_totals["Dispensing Fee Diff ($)"] += df["Dispensing Fee Diff ($)"].sum()

    # Convert the grand totals dictionary to a DataFrame
    grand_totals_df = pd.DataFrame([grand_totals])
    result_dict[period] = grand_totals_df

    return result_dict

def create_excel_from_dict(result_dict, period_start, period_end, summary_period):
    """
    Create an Excel file from a nested dictionary where DataFrame values are stored under
    primary (Unique_identificatr) and secondary ('brand', 'generic', 'current-period-total','year-to-date-total') keys.
    """
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Define styles
    bold_font = Font(bold=True)
    large_bold_font = Font(bold=True, size=14)
    left_align = Alignment(horizontal='left')
    right_align = Alignment(horizontal='right')
    center_align = Alignment(horizontal='center')
    light_grey_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Set up row index
    row_idx = 1
    header_written = False # need this flag in order to write column names only once

    for primary_key, sub_dict in result_dict.items():
        if primary_key in ['year-to-date-total','current-period-total']:
            continue

        # First row (empty)
        row_idx += 1
        col_span_max = max(
            len(sub_dict['brand'].columns) if 'brand' in sub_dict and not sub_dict['brand'].empty else 0,
            len(sub_dict['generic'].columns) if 'generic' in sub_dict and not sub_dict['generic'].empty else 0
        )
        # Iterate over sub dictionary (brand/generic)
        for secondary_key, df in sub_dict.items():
            if not df.empty:
                for col in df.columns:
                    if pd.api.types.is_categorical_dtype(df[col]):
                        df[col] = df[col].astype(str)  # Convert categorical to string
                        df[col].fillna("", inplace=True)  # Fill with empty string

                row_idx = write_rows(ws, df, row_idx, header_written, thin_border, center_align, bold_font)
                header_written = True 

    # Adding the total summary DataFrame at the end
    for total_summary in ['current-period-total','year-to-date-total']:
        if total_summary in result_dict:
            total_df = result_dict[total_summary]
            if total_summary == 'year-to-date-total':
                # Check the year of period_start and set the value based on the condition
                if period_start.year == 2024:
                    period_start = pd.to_datetime('2024-05-15') # Contract effective date
                else:
                    period_start = pd.to_datetime(f'{period_start.year}-01-01')
                period = 'Year to date'
            else:
                period = f'Total {summary_period}ly'
            if not total_df.empty:
                row_idx += 2  # Ensure there are two blank rows before the total summary
                col_span = len(total_df.columns)
                # Total summary header
                ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=col_span+1)
                cell = ws.cell(row=row_idx, column=1, value=f"{period} Summary for the period {period_start.strftime('%Y-%m-%d')} - {period_end.strftime('%Y-%m-%d')}")
                cell.alignment = left_align
                cell.font = large_bold_font
                cell.fill = light_grey_fill
                if total_summary == 'current-period-total':
                    cell = ws.cell(row=row_idx+2, column=1, value=f"Positive value mean Surplus; Negative value mean Deficiency -->")
                    cell.alignment = right_align
                    cell.font = bold_font
                
                # Total summary DataFrame
                row_idx += 1
                for r_idx, r in enumerate(dataframe_to_rows(total_df, index=False, header=True), start=row_idx):
                    for col_idx, value in enumerate(r, start=2):
                        cell = ws.cell(row=r_idx, column=col_idx, value=value)
                        cell.border = thin_border
                        cell.alignment = center_align
                        if r_idx == row_idx:  # Header row
                            cell.font = bold_font
                    row_idx = r_idx
                row_idx += 1  # Increment for the blank row after the total summary

        # Adjust column widths
        for col_idx in range(1, col_span_max + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for cell in ws[column_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Move the cursor to the beginning of the stream
    return output

def write_rows(ws, df, row_idx, header_written, thin_border, center_align, bold_font):
    """
    Write rows from a DataFrame to an Excel worksheet.
    
    Parameters:
    ws : openpyxl.Workbook.active : The active worksheet.
    df : pandas.DataFrame : The DataFrame to write.
    row_idx : int : The starting row index.
    header_written : bool : Whether the header has already been written.
    thin_border : openpyxl.styles.Border : The border style for cells.
    center_align : openpyxl.styles.Alignment : The alignment style for cells.
    bold_font : openpyxl.styles.Font : The font style for bold text.
    
    Returns:
    int : The updated row index after writing the DataFrame.
    """
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=not header_written), start=row_idx):
        for col_idx, value in enumerate(r, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = center_align
            if header_written and col_idx == 1 and r_idx != row_idx:  # First column, not header row
                cell.font = Font(bold=False)
            elif not header_written and r_idx == row_idx:  # Header row
                cell.font = bold_font
        row_idx = r_idx
    return row_idx
